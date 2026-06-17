#!/usr/bin/env python3
"""Generate the OpenO OLIS nomenclature seed CSVs from an official eHealth Ontario
OLIS Nomenclatures XLSX distribution.

The two seed files (``OLISTestResultNomenclature.csv`` and
``OLISTestRequestNomenclature.csv`` under ``database/mysql/olis/``) are the
install-time floor loaded by ``olisinit.sql`` via ``LOAD DATA LOCAL INFILE``.
At runtime the catalog is refreshed through Admin -> "OLIS - Import Nomenclature";
this script keeps the *seed* current so a fresh install starts close to the
live catalog instead of an older pinned version.

It mirrors ``OLISNomenclatureImportService`` exactly so the seed matches what the
Admin importer would produce from the same file:

  Result sheet  "Test Result Nomenclatures":
    nameId               <- "LOINC Code"            (row skipped if blank)
    name                 <- "Result Alternate Name 1"   (trimmed; "" if blank)
    status               <- deriveStatus(row)
    effectiveDate        <- "Effective Date"        (date or \\N)
    endDate              <- "End Date"              (date or \\N)
    externalCodeVersion  <- "External Code Version" (\\N if blank)
    sortKey              <- "Sort Key"              (\\N if blank)

  Request sheet "Test Request Nomenclature":
    nameId               <- "OLIS Test Request Code"     (row skipped if blank)
    name                 <- "Request Alternate Name 1"   (trimmed; "" if blank)
    category             <- "Test Request Category"      (\\N if blank)
    status               <- deriveStatus(row)
    effectiveDate, endDate, externalCodeVersion          (as above)
    sortKey              <- "Sort Key"              (\\N if blank)

The OLIS catalog sort key feeds the CV04/05/06/15 display ordering as the fallback
sort key when a result/request carries no in-message sort key (ZBX.2 / ZBR.11).

deriveStatus: INACTIVE if "Validation Status Indicator" == INACTIVE, or if
"Workflow Status Indicator" is present and != RELEASED; otherwise ACTIVE.

Output column order matches the ``LOAD DATA`` field lists in olisinit.sql
(sortKey appended last):
  result:  nameId, name, status, effectiveDate, endDate, externalCodeVersion, sortKey
  request: nameId, name, category, status, effectiveDate, endDate, externalCodeVersion, sortKey

Tab-delimited, '\\n' line terminator, NULL written as the MySQL ``\\N`` marker,
fields optionally double-quoted (matching ``OPTIONALLY ENCLOSED BY '"'``).

Usage:
    python3 scripts/olis/generate_nomenclature_seed.py \\
        --xlsx "OLIS Nomenclatures V3.04_PROD.xlsx" \\
        --out database/mysql/olis

Requires: openpyxl.
"""

import argparse
import csv
import datetime
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

RESULT_SHEET = "Test Result Nomenclatures"
REQUEST_SHEET = "Test Request Nomenclature"
NULL = r"\N"


def trim_to_null(value):
    """Mirror OLISNomenclatureImportService.trimToNull: blank -> None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def derive_status(row):
    """Mirror OLISNomenclatureImportService.deriveStatus."""
    validation = trim_to_null(row.get("Validation Status Indicator"))
    if validation is not None and validation.upper() == "INACTIVE":
        return "INACTIVE"
    workflow = trim_to_null(row.get("Workflow Status Indicator"))
    if workflow is not None and workflow.upper() != "RELEASED":
        return "INACTIVE"
    return "ACTIVE"


def fmt_date(value):
    """Format a date cell as YYYY-MM-DD, or \\N. openpyxl yields datetime for
    date-typed cells; strings/serials are parsed like parseDateCell."""
    if value is None or value == "":
        return NULL
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return NULL
    # Excel serial day number (days since 1899-12-30).
    try:
        serial = float(s)
        base = datetime.date(1899, 12, 30)
        return (base + datetime.timedelta(days=int(serial))).strftime("%Y-%m-%d")
    except ValueError:
        pass
    for pat in ("%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(s, pat).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return NULL


def read_rows(xlsx_path, sheet_name):
    """Yield each data row as a {header: value} dict. Row 1 is the header;
    only columns with a non-blank header are kept (mirrors OlisXlsxSheetReader)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        sys.exit("Sheet '%s' not found. Sheets: %s" % (sheet_name, wb.sheetnames))
    ws = wb[sheet_name]
    headers = None
    for cells in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [(str(c).strip() if c is not None else "") for c in cells]
            continue
        if all(c is None or str(c).strip() == "" for c in cells):
            continue
        row = {}
        for i, h in enumerate(headers):
            if h and i < len(cells):
                row[h] = cells[i]
        yield row
    wb.close()


def build_result_records(xlsx_path):
    seen, records = {}, []
    for row in read_rows(xlsx_path, RESULT_SHEET):
        name_id = trim_to_null(row.get("LOINC Code"))
        if name_id is None:
            continue
        rec = [
            name_id,
            (str(row.get("Result Alternate Name 1") or "")).strip(),
            derive_status(row),
            fmt_date(row.get("Effective Date")),
            fmt_date(row.get("End Date")),
            trim_to_null(row.get("External Code Version")) or NULL,
            trim_to_null(row.get("Sort Key")) or NULL,
        ]
        # upsert semantics: last occurrence of a nameId wins (as the importer merges)
        if name_id in seen:
            records[seen[name_id]] = rec
        else:
            seen[name_id] = len(records)
            records.append(rec)
    return records


def build_request_records(xlsx_path):
    seen, records = {}, []
    for row in read_rows(xlsx_path, REQUEST_SHEET):
        name_id = trim_to_null(row.get("OLIS Test Request Code"))
        if name_id is None:
            continue
        rec = [
            name_id,
            (str(row.get("Request Alternate Name 1") or "")).strip(),
            trim_to_null(row.get("Test Request Category")) or NULL,
            derive_status(row),
            fmt_date(row.get("Effective Date")),
            fmt_date(row.get("End Date")),
            trim_to_null(row.get("External Code Version")) or NULL,
            trim_to_null(row.get("Sort Key")) or NULL,
        ]
        if name_id in seen:
            records[seen[name_id]] = rec
        else:
            seen[name_id] = len(records)
            records.append(rec)
    return records


def write_csv(path, records):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, delimiter="\t", quotechar='"',
                       quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
        for rec in records:
            w.writerow(rec)


def summarize(label, records, status_idx):
    active = sum(1 for r in records if r[status_idx] == "ACTIVE")
    inactive = len(records) - active
    print("  %-8s %d rows  (ACTIVE %d, INACTIVE %d)" % (label, len(records), active, inactive))


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", required=True, help="Official OLIS Nomenclatures XLSX")
    ap.add_argument("--out", default="database/mysql/olis",
                    help="Output directory (default: database/mysql/olis)")
    args = ap.parse_args()

    if not os.path.isfile(args.xlsx):
        sys.exit("XLSX not found: %s" % args.xlsx)
    os.makedirs(args.out, exist_ok=True)

    print("Reading %s" % args.xlsx)
    results = build_result_records(args.xlsx)
    requests = build_request_records(args.xlsx)

    result_path = os.path.join(args.out, "OLISTestResultNomenclature.csv")
    request_path = os.path.join(args.out, "OLISTestRequestNomenclature.csv")
    write_csv(result_path, results)
    write_csv(request_path, requests)

    print("Wrote:")
    summarize("result", results, 2)   # status at index 2
    summarize("request", requests, 3)  # status at index 3
    print("  -> %s" % result_path)
    print("  -> %s" % request_path)
    print("\nRemember to update the version/date/source comments in olisinit.sql.")


if __name__ == "__main__":
    main()
