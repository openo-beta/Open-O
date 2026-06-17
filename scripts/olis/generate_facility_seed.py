#!/usr/bin/env python3
"""Generate the OpenO OLIS Lab/SCC facility seed CSV from an official eHealth
Ontario "Lab and SCC Extract" XLSX distribution.

The seed file (``OLISFacility.csv`` under ``database/mysql/olis/``) is the
install-time floor for the ``OLISFacility`` roster — the Lab and Specimen
Collection Centre directory that backs the facility typeahead pickers on OLIS
Search (Reporting/Performing/Ordering facility, etc.). At runtime the roster is
refreshed through Admin -> "OLIS - Import Lab/SCC"; this script keeps the *seed*
current so a fresh install starts with a populated roster instead of empty
dropdowns.

It mirrors ``OLISFacilityImportService.importRow`` exactly:

    licence              <- "Licence Number"            (row skipped if blank)
    oid                  <- "OID"                       (row skipped if blank)
    name                 <- "Facility Name"             (row skipped if blank)
    facilityClass        <- classFromOid(oid):
                              2.16.840.1.113883.3.59.1 -> LAB
                              2.16.840.1.113883.3.59.2 -> SCC
                              (anything else -> row skipped)
    addressLine1         <- "Facility Address Line One"
    addressLine2         <- "Facility Address Line Two"
    city                 <- "Facility Address City"
    postalCode           <- "Facility Address Postal_Code"
    status               <- "ACTIVE"   (the extract is the active roster)

Rows are upserted by (facilityClass, licence) in the importer, so duplicates are
collapsed last-wins here too.

Output column order matches the ``LOAD DATA`` field list to add to olisinit.sql:
    licenceNumber, facilityClass, name, addressLine1, addressLine2,
    city, postalCode, oid, status

Tab-delimited, '\\n' line terminator, NULL written as the MySQL ``\\N`` marker,
fields optionally double-quoted (matching ``OPTIONALLY ENCLOSED BY '"'``).

Usage:
    python3 scripts/olis/generate_facility_seed.py \\
        --xlsx "lab_and_scc_extract.xlsx" --out database/mysql/olis

Requires: openpyxl.
"""

import argparse
import csv
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

NULL = r"\N"
OID_LAB = "2.16.840.1.113883.3.59.1"
OID_SCC = "2.16.840.1.113883.3.59.2"
OID_HOSP = "2.16.840.1.113883.3.59.3"


def trim_to_null(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def class_from_oid(oid):
    if oid == OID_LAB:
        return "LAB"
    if oid == OID_SCC:
        return "SCC"
    if oid == OID_HOSP:
        return "HOS"
    return None


def build_hospital_records(xlsx_path):
    """Build HOS rows from an eHealth "Extract - Hospitals" XLSX (OID .59.3). The
    hospital extract uses a different header set than the Lab/SCC extract:

        licence       <- "Alternate Code"   (row skipped if blank)
        name          <- "OrgName"          (row skipped if blank)
        oid           <- "OID"              (must be the hospital OID .59.3)
        addressLine1  <- "Address Line 1"
        addressLine2  <- "Address Line 2"
        city          <- "City"
        postalCode    <- "Postal Code"
        status        <- "Status Code": OPEN -> ACTIVE, CLOSED -> INACTIVE

    Province is present in the extract but OLISFacility has no province column, so
    it is not loaded (facility addresses are otherwise sourced from the HL7 ZBR).
    Upsert key (facilityClass, licence); last occurrence wins.
    """
    seen, records = {}, []
    skipped = 0
    for row in read_rows(xlsx_path):
        licence = trim_to_null(row.get("Alternate Code"))
        oid = trim_to_null(row.get("OID"))
        name = trim_to_null(row.get("OrgName"))
        if licence is None or oid is None or name is None:
            skipped += 1
            continue
        facility_class = class_from_oid(oid)
        if facility_class != "HOS":
            skipped += 1
            continue
        status_code = (trim_to_null(row.get("Status Code")) or "").upper()
        status = "INACTIVE" if status_code == "CLOSED" else "ACTIVE"
        rec = [
            licence,
            facility_class,
            name,
            trim_to_null(row.get("Address Line 1")) or NULL,
            trim_to_null(row.get("Address Line 2")) or NULL,
            trim_to_null(row.get("City")) or NULL,
            trim_to_null(row.get("Postal Code")) or NULL,
            oid,
            status,
        ]
        key = (facility_class, licence)
        if key in seen:
            records[seen[key]] = rec
        else:
            seen[key] = len(records)
            records.append(rec)
    return records, skipped


def read_rows(xlsx_path):
    """Yield each data row as a {header: value} dict; row 1 is the header
    (mirrors OlisXlsxSheetReader: only non-blank header columns are kept)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  # importer uses firstSheetPath
    headers = None
    for cells in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [(str(c).strip() if c is not None else "") for c in cells]
            continue
        if all(c is None or str(c).strip() == "" for c in cells):
            continue
        row = {}
        for i, h in enumerate(headers):
            if h and i < len(cells):
                row[h] = cells[i]
        yield row
    wb.close()


def build_records(xlsx_path):
    seen, records = {}, []
    skipped = 0
    for row in read_rows(xlsx_path):
        licence = trim_to_null(row.get("Licence Number"))
        oid = trim_to_null(row.get("OID"))
        name = trim_to_null(row.get("Facility Name"))
        if licence is None or oid is None or name is None:
            skipped += 1
            continue
        facility_class = class_from_oid(oid)
        if facility_class is None:
            skipped += 1
            continue
        rec = [
            licence,
            facility_class,
            name,
            trim_to_null(row.get("Facility Address Line One")) or NULL,
            trim_to_null(row.get("Facility Address Line Two")) or NULL,
            trim_to_null(row.get("Facility Address City")) or NULL,
            trim_to_null(row.get("Facility Address Postal_Code")) or NULL,
            oid,
            "ACTIVE",
        ]
        key = (facility_class, licence)  # upsert key
        if key in seen:
            records[seen[key]] = rec
        else:
            seen[key] = len(records)
            records.append(rec)
    return records, skipped


def write_csv(path, records):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, delimiter="\t", quotechar='"',
                       quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
        for rec in records:
            w.writerow(rec)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", required=True, help="Official eHealth Lab and SCC Extract XLSX")
    ap.add_argument("--hospital-xlsx", default=None,
                    help="Optional eHealth 'Extract - Hospitals' XLSX (OID .59.3); appends HOS rows")
    ap.add_argument("--out", default="database/mysql/olis",
                    help="Output directory (default: database/mysql/olis)")
    args = ap.parse_args()

    if not os.path.isfile(args.xlsx):
        sys.exit("XLSX not found: %s" % args.xlsx)
    os.makedirs(args.out, exist_ok=True)

    print("Reading %s" % args.xlsx)
    records, skipped = build_records(args.xlsx)

    hosp_count = 0
    if args.hospital_xlsx:
        if not os.path.isfile(args.hospital_xlsx):
            sys.exit("Hospital XLSX not found: %s" % args.hospital_xlsx)
        print("Reading %s" % args.hospital_xlsx)
        hosp_records, hosp_skipped = build_hospital_records(args.hospital_xlsx)
        records.extend(hosp_records)
        hosp_count = len(hosp_records)
        skipped += hosp_skipped

    labs = sum(1 for r in records if r[1] == "LAB")
    sccs = sum(1 for r in records if r[1] == "SCC")
    hos = sum(1 for r in records if r[1] == "HOS")

    out_path = os.path.join(args.out, "OLISFacility.csv")
    write_csv(out_path, records)

    print("Wrote %d rows (LAB %d, SCC %d, HOS %d); skipped %d (missing licence/oid/name or unknown OID)"
          % (len(records), labs, sccs, hos, skipped))
    print("  -> %s" % out_path)
    print("\nAdd a LOAD DATA stanza for OLISFacility to olisinit.sql (see field order above).")


if __name__ == "__main__":
    main()
